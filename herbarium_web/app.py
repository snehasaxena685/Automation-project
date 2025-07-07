from flask import Flask, render_template, request, redirect, url_for, send_from_directory
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import os, re, shutil, requests

EXCEL = "Live_Herbarium_Log.xlsx"
IMAGEDIR = "static/images"
CSV = "herbs_data.csv"
API_KEY = os.getenv("PERENUAL_API_KEY", "")

os.makedirs(IMAGEDIR, exist_ok=True)
app = Flask(__name__)

def get_next_id():
    if os.path.exists(EXCEL):
        df = pd.read_excel(EXCEL)
        if "Specimen ID" in df and not df.empty:
            last = df["Specimen ID"].dropna().iloc[-1]
            num = int(re.search(r"HSPC(\d+)", str(last)).group(1)) + 1
            return f"HSPC{num:03d}"
    return "HSPC001"

def fetch_api(name):
    try:
        r = requests.get(f"https://perenual.com/api/species-list?key={API_KEY}&q={name}", timeout=5).json()
        d = r.get("data", [{}])[0]
        return d.get("scientific_name",""), d.get("default_image",{}).get("original_url","")
    except: return "", ""

@app.route("/", methods=["GET", "POST"])
def home():
    if request.method == "POST":
        sid = request.form["sid"]
        herb = request.form["herb"].title()
        sci = request.form["sci"]
        region = request.form["region"]
        weight = request.form["weight"]
        img_url = request.form.get("imgurl","")
        date = datetime.today().strftime("%Y-%m-%d")
        wb = load_workbook(EXCEL) if os.path.exists(EXCEL) else load_workbook(load_workbook=[])
        ws = wb.active
        if ws.max_column < 8: ws.append(["Specimen ID","Herb Name","Scientific Name","Date","Region","Weight","Image_URL","Image_File"])
        # Download image:
        img_file = ""
        if img_url:
            ext = os.path.splitext(img_url)[1]
            img_file = f"{sid}{ext}"
            try:
                r = requests.get(img_url, timeout=5)
                with open(os.path.join(IMAGEDIR, img_file), "wb") as f: f.write(r.content)
            except: img_file = ""
        ws.append([sid, herb, sci, date, region, weight, img_url, img_file])
        wb.save(EXCEL)
        return redirect(url_for("home"))

    sid = get_next_id()
    df = pd.read_csv(CSV)
    herbs = sorted(df["Herb Name"].dropna().unique())
    return render_template("index.html", sid=sid, herbs=herbs)

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.getenv("PORT",5000)))

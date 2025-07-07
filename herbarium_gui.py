import tkinter as tk
from tkinter import ttk, messagebox, Toplevel, filedialog
from openpyxl import load_workbook
from datetime import datetime
from PIL import Image, ImageTk
import pandas as pd
import os
import re
import shutil
import requests

# === CONFIGURATION ===
EXCEL_PATH = "Live_Herbarium_Log.xlsx"
IMAGE_DIR = "herb_images"
API_KEY = "sk-8m6Q686a51d0266f311312"  # Replace this
CSV_PATH = "herbs_data.csv"

selected_image_path = None
os.makedirs(IMAGE_DIR, exist_ok=True)

def get_next_specimen_id():
    try:
        df = pd.read_excel(EXCEL_PATH)
        if "Specimen ID" not in df.columns or df.empty:
            return "HSPC001"
        last_id = df["Specimen ID"].dropna().iloc[-1]
        match = re.search(r"HSPC(\d+)", str(last_id).upper())
        if match:
            return f"HSPC{int(match.group(1)) + 1:03d}"
    except:
        pass
    return "HSPC001"

def classify_priority(region):
    return "High" if region and any(r in region for r in ["Delhi", "Punjab", "Uttar Pradesh"]) else "Low"

def fetch_from_api(herb):
    try:
        url = f"https://perenual.com/api/species-list?key={API_KEY}&q={herb}"
        response = requests.get(url, timeout=10)
        json_data = response.json()

        if not isinstance(json_data, dict) or "data" not in json_data:
            raise ValueError("Invalid response structure.")
        data = json_data.get("data", [])
        if data:
            plant = data[0]
            sci_name = plant.get("scientific_name", "Not Found")
            img_url = plant.get("default_image", {}).get("original_url", "")
            return sci_name, "India", 20, img_url
        else:
            raise ValueError("No plant data found.")
    except Exception as e:
        print(f"API Error: {e}")
        return "Not Found Online", "", "", ""

def auto_fill_fields(event=None):
    herb_name = entry_herb.get().strip().title()
    if not herb_name:
        return
    try:
        df = pd.read_csv(CSV_PATH)
        match = df[df["Herb Name"].str.lower() == herb_name.lower()]
        if not match.empty:
            sci_name = match.iloc[0]["Scientific Name"]
            region = match.iloc[0]["Region"]
            weight = match.iloc[0]["Avg Weight (g)"]
            img_url = ""
        else:
            sci_name, region, weight, img_url = fetch_from_api(herb_name)

        entry_sci.delete(0, tk.END)
        entry_sci.insert(0, sci_name)
        combo_region.set(region)
        entry_weight.delete(0, tk.END)
        entry_weight.insert(0, weight)
        lbl_api_img.config(text=f"{'🌐 Found Online' if img_url else '❌ Not Found'}", fg="green" if img_url else "red")
    except Exception as e:
        print(f"Auto-fill Error: {e}")

def upload_image():
    global selected_image_path
    file_path = filedialog.askopenfilename(filetypes=[("Image Files", "*.png *.jpg *.jpeg")])
    if file_path:
        selected_image_path = file_path
        lbl_img.config(text=os.path.basename(file_path), fg="green")

def submit_data():
    global selected_image_path
    sid = entry_id.get()
    herb = entry_herb.get()
    sci_name = entry_sci.get()
    region = combo_region.get()
    weight = entry_weight.get()
    date = datetime.today().strftime("%Y-%m-%d")

    if not sid or not herb or not sci_name or not region or not weight:
        messagebox.showerror("❌ Incomplete", "Please fill all fields.")
        return

    try:
        weight = float(weight)
    except:
        messagebox.showerror("❌ Invalid", "Weight must be a number.")
        return

    priority = classify_priority(region)
    image_filename = ""

    if selected_image_path:
        ext = os.path.splitext(selected_image_path)[1]
        image_filename = f"{sid}_{herb.lower().replace(' ', '_')}{ext}"
        dest_path = os.path.join(IMAGE_DIR, image_filename)
        shutil.copy2(selected_image_path, dest_path)

    try:
        wb = load_workbook(EXCEL_PATH)
        sheet = wb.active
        if sheet.max_column < 8:
            sheet.append(["Specimen ID", "Herb Name", "Scientific Name", "Date", "Region", "Weight", "Priority", "Image"])
        sheet.append([sid, herb, sci_name, date, region, weight, priority, image_filename])
        wb.save(EXCEL_PATH)
        messagebox.showinfo("✅ Saved", "Entry saved successfully!")
    except PermissionError:
        messagebox.showerror("🔒 File Locked", "Close the Excel file first.")
        return

    entry_id.delete(0, tk.END)
    entry_id.insert(0, get_next_specimen_id())
    entry_herb.set('')
    entry_sci.delete(0, tk.END)
    combo_region.set('')
    entry_weight.delete(0, tk.END)
    lbl_img.config(text="No image selected", fg="grey")
    lbl_api_img.config(text="", fg="grey")
    selected_image_path = None

def open_viewer_popup():
    def load_data():
        try:
            return pd.read_excel(EXCEL_PATH)
        except Exception as e:
            messagebox.showerror("Error", f"Could not load data: {e}")
            return pd.DataFrame()

    def apply_filters():
        df = load_data()
        herb = herb_filter.get().strip().lower()
        region = region_filter.get()
        priority = priority_filter.get()
        if herb:
            df = df[df['Herb Name'].str.lower().str.contains(herb)]
        if region:
            df = df[df['Region'] == region]
        if priority:
            df = df[df['Priority'] == priority]
        update_table(df)

    def update_table(df):
        for row in tree.get_children():
            tree.delete(row)
        for _, row in df.iterrows():
            tree.insert('', tk.END, values=list(row))

    viewer = Toplevel()
    viewer.title("🔍 View & Filter Entries")
    viewer.geometry("960x520")
    viewer.configure(bg="#f1f8fc")

    filter_frame = tk.Frame(viewer, bg="#f1f8fc", pady=10)
    filter_frame.pack(fill=tk.X)

    tk.Label(filter_frame, text="Herb Name:", bg="#f1f8fc").pack(side=tk.LEFT, padx=5)
    herb_filter = tk.Entry(filter_frame)
    herb_filter.pack(side=tk.LEFT, padx=5)

    tk.Label(filter_frame, text="Region:", bg="#f1f8fc").pack(side=tk.LEFT, padx=5)
    region_filter = ttk.Combobox(filter_frame, values=[
        "", "Delhi", "Punjab", "Uttar Pradesh", "Gujarat", "Maharashtra",
        "Rajasthan", "Karnataka", "West Bengal", "Haryana", "Madhya Pradesh"
    ])
    region_filter.pack(side=tk.LEFT, padx=5)

    tk.Label(filter_frame, text="Priority:", bg="#f1f8fc").pack(side=tk.LEFT, padx=5)
    priority_filter = ttk.Combobox(filter_frame, values=["", "High", "Low"])
    priority_filter.pack(side=tk.LEFT, padx=5)

    tk.Button(filter_frame, text="Apply Filters", command=apply_filters,
              bg="#007B8A", fg="white").pack(side=tk.LEFT, padx=10)

    columns = ["Specimen ID", "Herb Name", "Scientific Name", "Date", "Region", "Weight", "Priority", "Image"]
    tree = ttk.Treeview(viewer, columns=columns, show='headings', height=18)
    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, width=120, anchor="center")
    tree.pack(fill=tk.BOTH, expand=True)

    update_table(load_data())

# === GUI SETUP ===
root = tk.Tk()
root.title("🌿 CSIR Herbarium Logger")
root.geometry("600x770")
root.configure(bg="#f0f8ff")

header = tk.Frame(root, bg="#005F6B", pady=10)
header.pack(fill=tk.X)

try:
    logo_img = Image.open("csir_logo.png").resize((60, 60))
    logo = ImageTk.PhotoImage(logo_img)
    tk.Label(header, image=logo, bg="#005F6B").pack()
except:
    tk.Label(header, text="[CSIR LOGO]", font=("Segoe UI", 16), bg="#005F6B", fg="white").pack()

tk.Label(header, text="CSIR Herbarium Digitization Tool", font=("Segoe UI", 16, "bold"),
         bg="#005F6B", fg="white").pack()

card = tk.Frame(root, bg="white", padx=25, pady=25, relief="groove", bd=2)
card.pack(padx=20, pady=20, fill=tk.BOTH, expand=True)

font_label = ("Segoe UI", 10)
font_input = ("Segoe UI", 11)

def add_field(label, widget):
    tk.Label(card, text=label, font=font_label, bg="white").pack(anchor="w", pady=(10, 2))
    widget.pack(fill="x", ipady=6)

entry_id = tk.Entry(card, font=font_input)
add_field("🔖 Specimen ID", entry_id)

try:
    herb_df = pd.read_csv(CSV_PATH)
    herb_list = sorted(herb_df["Herb Name"].dropna().unique().tolist())
except:
    herb_list = []

entry_herb = ttk.Combobox(card, values=herb_list, font=font_input)
entry_herb.bind("<<ComboboxSelected>>", auto_fill_fields)
entry_herb.bind("<FocusOut>", auto_fill_fields)
add_field("🌿 Herb Name", entry_herb)

entry_sci = tk.Entry(card, font=font_input)
add_field("🔬 Scientific Name", entry_sci)

combo_region = ttk.Combobox(card, font=font_input, state="readonly")
combo_region['values'] = [
    "Delhi", "Punjab", "Uttar Pradesh", "Gujarat", "Maharashtra",
    "Rajasthan", "Karnataka", "West Bengal", "Haryana", "Madhya Pradesh"
]
add_field("📍 Region", combo_region)

entry_weight = tk.Entry(card, font=font_input)
add_field("⚖️ Weight (grams)", entry_weight)

tk.Label(card, text="🌐 Online Data Status", font=font_label, bg="white").pack(anchor="w")
lbl_api_img = tk.Label(card, text="", font=("Segoe UI", 9), fg="grey", bg="white")
lbl_api_img.pack(anchor="w")

tk.Label(card, text="📸 Upload Herb Image", font=font_label, bg="white").pack(anchor="w", pady=(10, 2))
tk.Button(card, text="Choose Image", command=upload_image, bg="#6c757d", fg="white").pack(fill="x", ipady=5)
lbl_img = tk.Label(card, text="No image selected", bg="white", fg="grey", font=("Segoe UI", 9))
lbl_img.pack(anchor="w", pady=(4, 8))

tk.Button(root, text="✅ Submit Entry", font=("Segoe UI", 11, "bold"),
          bg="#198754", fg="white", activebackground="#146c43",
          padx=14, pady=8, command=submit_data).pack(pady=(5, 6))

tk.Button(root, text="📂 View Entries", font=("Segoe UI", 11),
          bg="#0d6efd", fg="white", activebackground="#0b5ed7",
          command=open_viewer_popup).pack(pady=(0, 20))

entry_id.insert(0, get_next_specimen_id())
root.mainloop()
